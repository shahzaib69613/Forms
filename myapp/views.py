
from django.shortcuts import render, redirect
from .forms import ContactForm
from django.shortcuts import render, redirect, get_object_or_404
from .models import Contact  # Your Contact model
from .forms import ContactForm  # Your ContactForm
from .forms import LoginForm
from .models import Authenticate as authenticate
from django.contrib.auth.decorators import login_required


@login_required
def contact_view(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            form.save()  # Save data to the database
            return render(request, 'thank_you.html', {'name': form.cleaned_data['name']})
    else:
        form = ContactForm()
    return render(request, 'contact_form.html', {'form': form})

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Contact
import json

@csrf_exempt  # For simplicity, exempt CSRF validation in this example
def check_retailer_id(request):
    retailer_id = request.GET.get('retailer_id')
    if retailer_id and Contact.objects.filter(Retailer_ID=retailer_id).exists():
        return JsonResponse({'exists': True})
    return JsonResponse({'exists': False})



from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.timezone import now
from .models import ContactChangeLog
import pytz
from django.utils.timezone import now

@login_required
def contact_edit_view(request, pk):
    contact = get_object_or_404(Contact, pk=pk)

    if request.method == 'POST':
        form = ContactForm(request.POST, instance=contact)

        if form.is_valid():
            # Fetch the original instance from the database again to ensure consistency
            original_contact = Contact.objects.get(pk=pk)
            
            # Track changes dynamically
            changes = {}

            # Get the dynamically changed fields from the form
            for field in form.changed_data:
                old_value = getattr(original_contact, field)  # Old value from the database
                new_value = form.cleaned_data[field]  # New value from the form

                # Add the change to the dictionary
                changes[field] = {
                    "old": old_value,
                    "new": new_value,
                }

            # Save the updated contact
            form.save()

            # Convert the timestamp to Pakistan Standard Time (PST)
            pakistan_tz = pytz.timezone('Asia/Karachi')
            timestamp = now().astimezone(pakistan_tz)

            # Log changes dynamically
            for field, change in changes.items():
                ContactChangeLog.objects.create(
                    username=request.user.username,
                    contact=contact,
                    field_name=field,
                    old_value=change["old"],
                    new_value=change["new"],
                    timestamp=timestamp,
                )

            return redirect('contact_edit', pk=pk)

    else:
        form = ContactForm(instance=contact)

    # Pass the original field values to the template
    initial_values = {field.name: getattr(contact, field.name) for field in form}

    return render(request, 'edit_contact.html', {'form': form, 'contact': contact, 'initial_values': initial_values})

from django.shortcuts import redirect

# @login_required
# def contact_list_view(request):
#     if request.method == 'POST':
#         form = ContactForm(request.POST)
#         if form.is_valid():
#             form.save()  # Save data to the database
#             # Redirect to the same page after saving the form data
#             return redirect('contact_list')  # Replace 'contact_list' with your URL name for the contact list view
        
#     else:
#         form = ContactForm()

#     contacts = Contact.objects.all()  # Fetch all records from the database
#     return render(request, 'contact_list.html', {'form': form, 'contacts': contacts})
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import Contact
from .forms import ContactForm

@login_required
def contact_list_view(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('contact_list')

    else:
        form = ContactForm()

    # Filter contacts where Franchise_ID matches the logged-in user's username
    contacts = Contact.objects.filter(Category='Retailer', Franchise_ID=request.user.username)
    
    return render(request, 'contact_list.html', {'form': form, 'contacts': contacts})


from django.shortcuts import render, redirect
from .forms import ContactForm
from .models import Contact
from django.contrib.auth.decorators import login_required

@login_required
def DSO_list_view(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            form.save()  # Save data to the database
            # Redirect to the same page after saving the form data
            return redirect('DSO_List')  # Replace 'contact_list' with your URL name for the contact list view
    else:
        form = ContactForm()

    # Apply the filter where Category is 'DSO'
    contacts = Contact.objects.filter(Category='DSO',Franchise_ID=request.user.username)  # Only fetch records where Category = 'DSO'

    return render(request, 'DSO_List.html', {'form': form, 'contacts': contacts})



@login_required
def RSO_list_view(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            form.save()  # Save data to the database
            # Redirect to the same page after saving the form data
            return redirect('RSO_list')  # Replace 'contact_list' with your URL name for the contact list view
         
    else:
        form = ContactForm()

    contacts = Contact.objects.filter(Category='RSO',Franchise_ID=request.user.username) # Fetch all records from the database
    return render(request, 'RSO_list.html', {'form': form, 'contacts': contacts})   


from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from .forms import LoginForm
from django.contrib.auth.models import User

import firebase_admin
from firebase_admin import auth, credentials
from django.http import JsonResponse
from django.middleware.csrf import get_token
from django.contrib.auth import login
from django.contrib.auth.models import User




import firebase_admin
from firebase_admin import credentials, firestore
import os
import base64
import json

# Get the FIREBASE_CREDENTIALS environment variable
firebase_creds_base64 = os.getenv("FIREBASE_CREDENTIALS")

if not firebase_creds_base64:
    raise ValueError("FIREBASE_CREDENTIALS environment variable is not set or empty.")

try:
    firebase_creds = json.loads(base64.b64decode(firebase_creds_base64).decode())
except json.JSONDecodeError:
    raise ValueError("Invalid FIREBASE_CREDENTIALS: Unable to decode JSON.")

# Initialize Firebase only if not already initialized
if not firebase_admin._apps:
    cred = credentials.Certificate(firebase_creds)
    firebase_admin.initialize_app(cred)


db = firestore.client() 
# Function to create test users in Firebase
def create_test_users():
    users_data = [
        {"username": "user1", "email": "user1@example.com", "password": "password123"},
        {"username": "user2", "email": "user2@example.com", "password": "password123"},
        {"username": "user3", "email": "user3@example.com", "password": "password123"},
    ]

    for data in users_data:
        try:
            # Create user in Firebase Authentication
            user = auth.create_user(email=data["email"], password=data["password"])
            
            # Store username mapping in Firestore
            db.collection("users").document(data["username"]).set({"email": data["email"]})

            print(f"User {data['username']} created successfully in Firebase.")
        except Exception as e:
            print(f"Error creating user {data['username']}: {e}")

# Run the function to create test users
create_test_users()
import firebase_admin
from firebase_admin import auth, credentials
from django.http import JsonResponse
from django.middleware.csrf import get_token
from django.contrib.auth import login
from django.contrib.auth.models import User


from firebase_admin import auth
from django.http import JsonResponse
from django.contrib.auth import login
from django.contrib.auth.models import User
import json

def firebase_login_view(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            id_token = data.get("token")

            decoded_token = auth.verify_id_token(id_token)
            email = decoded_token.get("email")

            if not email:
                return JsonResponse({"success": False, "message": "Email not found in token"}, status=400)

            user, created = User.objects.get_or_create(username=email, defaults={"email": email})

            login(request, user)

            return JsonResponse({"success": True, "message": "Login successful"})
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)}, status=400)

    return JsonResponse({"success": False, "message": "Invalid request method"}, status=405)

def login_view(request):
    return render(request, "login.html") 

# views.py
import io
import csv
from django.shortcuts import render
from django.http import HttpResponse
from .forms import FileUploadForm
from .models import DataRecord

import io
import csv
from django.http import HttpResponse
from django.shortcuts import render
from .forms import FileUploadForm
from .models import DataRecord




from django.http import JsonResponse
from .models import DataRecord

def get_retailer_number(request):
    retailer_id = request.GET.get('retailer_id')
    if retailer_id:
        try:
            data_record = DataRecord.objects.get(retailer_id=retailer_id)
           
            return JsonResponse({'retailer_number': data_record.retailer_msisdn,'Franchise_ID': data_record.second_parent_id,
            'Region': data_record.second_parent_region})
        except DataRecord.DoesNotExist:
            return JsonResponse({'retailer_number': None})
    return JsonResponse({'retailer_number': None})

import io
import csv
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import load_workbook
from .forms import FileUploadForm, BVSUploadForm, HeirarchyUploadForm
from .models import DataRecord, DataRecordBVS, Heirarchy

def upload_file(request):
    form1 = FileUploadForm()
    form2 = BVSUploadForm()
    form3 = HeirarchyUploadForm()

    if request.method == 'POST':
        if 'upload_csv' in request.POST:
            form1 = FileUploadForm(request.POST, request.FILES)
            if form1.is_valid():
                csv_file = request.FILES['file']
                decoded_file = csv_file.read().decode('utf-8')
                io_string = io.StringIO(decoded_file)
                csv_reader = csv.reader(io_string, delimiter=',', quotechar='"')

                # Clear all existing records for DataRecord
                DataRecord.objects.all().delete()

                # Collect data for bulk creation
                records = [
                    DataRecord(
                        second_parent_region=row[0],
                        second_parent_id=row[1],
                        rso_id=row[2],
                        rso_msisdn=row[3],
                        retailer_id=row[4],
                        retailer_msisdn=row[5]
                    )
                    for row in csv_reader
                ]

                DataRecord.objects.bulk_create(records)
                return HttpResponse('CSV file for DataRecord uploaded and data inserted successfully.')

        elif 'upload_bvs_xlsx' in request.POST:
            form2 = BVSUploadForm(request.POST, request.FILES)
            if form2.is_valid():
                excel_file = request.FILES['file']
                wb = load_workbook(excel_file)
                sheet = wb.active

                # Clear all existing records for DataRecordBVS
                DataRecordBVS.objects.all().delete()

                records = [
                    DataRecordBVS(Device_ID=row[0], retailer_id=row[2])
                    for row in sheet.iter_rows(min_row=2, values_only=True)
                ]

                DataRecordBVS.objects.bulk_create(records)
                return HttpResponse('XLSX file for DataRecordBVS uploaded and data inserted successfully.')

        elif 'upload_heirarchy_xlsx' in request.POST:
            form3 = HeirarchyUploadForm(request.POST, request.FILES)
            if form3.is_valid():
                excel_file = request.FILES['file']
                wb = load_workbook(excel_file)
                sheet = wb["HQ"]  # Explicitly select the "HQ" sheet

                # Clear all existing records for Heirarchy
                Heirarchy.objects.all().delete()

                records = [
                    Heirarchy(Franchise_ID=row[0], Grid=row[10],Cluster=row[23])
                    for row in sheet.iter_rows(min_row=3, values_only=True)
                ]

                Heirarchy.objects.bulk_create(records)
                return HttpResponse('XLSX file for Heirarchy uploaded and data inserted successfully.')

    return render(request, 'upload.html', {'form1': form1, 'form2': form2, 'form3': form3})

from .models import DataRecordBVS
from django.http import JsonResponse
from .models import DataRecordBVS,Heirarchy

from django.http import JsonResponse
from myapp.models import Heirarchy  # Import your model

def get_Cluster(request):
    Franchise_ID = request.GET.get('Franchise_ID')
    if Franchise_ID:
        try:
            # Fetch data from the Heirarchy model
            data_record = Heirarchy.objects.get(Franchise_ID=Franchise_ID)
            print(data_record.Grid)  # Debugging output to verify the value
            return JsonResponse({'Grid': data_record.Grid, 'Cluster': data_record.Cluster})
        except Heirarchy.DoesNotExist:  # Handle the exception properly
            # Return None if the record doesn't exist
            return JsonResponse({'Grid': None, 'Cluster': None}, status=404)
    # Return None if no Franchise_ID is provided
    return JsonResponse({'Grid': None, 'Cluster': None}, status=400)


def get_bvs_number(request):
    retailer_id = request.GET.get('retailer_id')
    if retailer_id:
        try:
            # Fetch data from the DataRecordBVS model
            data_record = DataRecordBVS.objects.get(retailer_id=retailer_id)
            print(data_record.Device_ID)
            return JsonResponse({'Device_ID': data_record.Device_ID})
        except DataRecordBVS.DoesNotExist:
            # Return None if the record doesn't exist
            return JsonResponse({'Device_ID': None})
    # Return None if no retailer_id is provided
    return JsonResponse({'Device_ID': None})

# In your views.py
from django.shortcuts import render, get_object_or_404, redirect
from .models import Contact

from django.shortcuts import get_object_or_404, redirect
from django.http import HttpResponse

from django.shortcuts import render, get_object_or_404, redirect
from .models import Contact
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse

from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
import json

@login_required
def contact_delete(request, contact_id):  
    print("Delete view called with ID:", contact_id)  # Debugging

    contact = get_object_or_404(Contact, id=contact_id)

    if request.method == 'POST':
        print("Request method is POST")  # Debugging
        contact.delete()
        return JsonResponse({"message": "Contact deleted successfully"}, status=200)

    print("Invalid request method:", request.method)  # Debugging
    return JsonResponse({"error": "Invalid request"}, status=400)

from django.shortcuts import render
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from .models import Contact  # Import your contact model

def generate_oath_pdf(request, contact_id):
    # Retrieve the contact from the database using the contact ID
    contact = Contact.objects.get(id=contact_id)
    
    # Create the PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{contact.DSO_Name}_Oath_Form.pdf"'
    
    # Create a canvas object to generate the PDF
    pdf = canvas.Canvas(response, pagesize=letter)
    
    # Leave half the page blank
    pdf.drawString(100, 100, "")
    
    # Set up the oath statement
    pdf.setFont("Helvetica", 12)
    pdf.drawString(80, 250, "I, _____, S/O _____, work at Zong Franchise __________________ as a ________________.")
    pdf.drawString(80, 230, f"The BVS device assigned to me has the IMEI number {contact.BVS_Device}.")
    pdf.drawString(80, 210, "I understand that I will be held responsible for any unlawful activity involving this device,")
    pdf.drawString(80, 190, "and the company reserves the right to take legal action against me.")
    
    # Add signature line
    pdf.drawString(80, 150, "Signature: ____________________________")
    pdf.drawString(80, 130, "Date: ________________________________")
    
    # Save the PDF
    pdf.showPage()
    pdf.save()
    
    return response
